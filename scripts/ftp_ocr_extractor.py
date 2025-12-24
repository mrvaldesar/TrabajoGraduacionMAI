import os
import ftplib
import random
import fitz  # PyMuPDF
from PIL import Image
import pytesseract
import shutil
from docx import Document
from openpyxl import load_workbook
import xlrd

# --- Validación de instalación de Tesseract OCR ---
try:
    pytesseract.get_tesseract_version()
    print("✅ Tesseract OCR detectado correctamente.")
except Exception:
    print("❌ ERROR: Tesseract OCR no está instalado o no está en el PATH.")
    print("Descárgalo desde: https://github.com/UB-Mannheim/tesseract/wiki")
    exit()

# --- Variables de Configuración ---

SOURCE_TYPE = "LOCAL"  # "FTP" o "LOCAL"

FTP_HOST = "ftp.example.com"
FTP_USER = "user"
FTP_PASS = "password"
FTP_REMOTE_DIR = "/remote/path"

LOCAL_SOURCE_DIR = "E:\\Documentos\\proyectos\\descargaPDF"

DESTINATION_DIR = "extracted_files"
FILE_LIMIT = 0
RANDOM_SELECTION = False

IMAGE_EXTENSIONS = ('.jpg', '.jpeg', '.png', '.tiff', '.bmp', '.webp')
DOC_EXTENSIONS = ('.docx',)
EXCEL_EXTENSIONS = ('.xls', '.xlsx')

# --- FTP ---

def connect_ftp():
    try:
        ftp = ftplib.FTP(FTP_HOST, timeout=30)
        ftp.login(user=FTP_USER, passwd=FTP_PASS)
        ftp.set_pasv(True)
        return ftp
    except ftplib.all_errors as e:
        print(f"Error FTP: {e}")
        return None

# --- Listado de archivos ---

def list_files_local(local_dir):
    files_found = []
    for root, _, files in os.walk(local_dir):
        for file in files:
            full_path = os.path.join(root, file)
            if file.lower().endswith(('.pdf',) + IMAGE_EXTENSIONS + DOC_EXTENSIONS + EXCEL_EXTENSIONS):
                print(f"Archivo encontrado: {full_path}")
                files_found.append(full_path)
    return files_found

# --- OCR y extracción ---

def ocr_pdf(pdf_path, output_txt_path):
    try:
        pdf_document = fitz.open(pdf_path)
        full_text = ""

        for page_num in range(len(pdf_document)):
            page = pdf_document.load_page(page_num)
            pix = page.get_pixmap(dpi=300)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            text = pytesseract.image_to_string(img)
            full_text += text + "\n\n"

        with open(output_txt_path, 'w', encoding='utf-8') as f:
            f.write(full_text)

        print(f"OCR PDF guardado: {output_txt_path}")
        return True
    except Exception as e:
        print(f"Error OCR PDF: {e}")
        return False

def ocr_image(image_path, output_txt_path):
    try:
        img = Image.open(image_path)
        text = pytesseract.image_to_string(img)

        with open(output_txt_path, 'w', encoding='utf-8') as f:
            f.write(text)

        print(f"OCR IMAGEN guardado: {output_txt_path}")
        return True
    except Exception as e:
        print(f"Error OCR IMAGEN: {e}")
        return False

def extract_text_from_docx(docx_path, output_txt_path):
    try:
        doc = Document(docx_path)
        full_text = "\n".join([p.text for p in doc.paragraphs])

        with open(output_txt_path, 'w', encoding='utf-8') as f:
            f.write(full_text)

        print(f"TEXTO WORD guardado: {output_txt_path}")
        return True
    except Exception as e:
        print(f"Error Word: {e}")
        return False

def extract_text_from_xlsx(xlsx_path, output_txt_path):
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        full_text = ""

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            full_text += f"\n--- Hoja: {sheet} ---\n"
            for row in ws.iter_rows(values_only=True):
                full_text += " | ".join([str(c) if c else "" for c in row]) + "\n"

        with open(output_txt_path, 'w', encoding='utf-8') as f:
            f.write(full_text)

        print(f"TEXTO XLSX guardado: {output_txt_path}")
        return True
    except Exception as e:
        print(f"Error XLSX: {e}")
        return False

def extract_text_from_xls(xls_path, output_txt_path):
    try:
        wb = xlrd.open_workbook(xls_path)
        full_text = ""

        for sheet in wb.sheets():
            full_text += f"\n--- Hoja: {sheet.name} ---\n"
            for r in range(sheet.nrows):
                row = sheet.row(r)
                full_text += " | ".join([str(c.value) for c in row]) + "\n"

        with open(output_txt_path, 'w', encoding='utf-8') as f:
            f.write(full_text)

        print(f"TEXTO XLS guardado: {output_txt_path}")
        return True
    except Exception as e:
        print(f"Error XLS: {e}")
        return False

# --- Procesamiento ---

def process_files(file_list, destination_dir, source_base_dir):
    for source_path in file_list:
        relative_path = os.path.relpath(source_path, source_base_dir)
        local_destination_path = os.path.join(destination_dir, relative_path)

        os.makedirs(os.path.dirname(local_destination_path), exist_ok=True)
        shutil.copy(source_path, local_destination_path)

        output_txt_path = os.path.splitext(local_destination_path)[0] + '.txt'
        ext = os.path.splitext(local_destination_path)[1].lower()

        if ext == '.pdf':
            ocr_pdf(local_destination_path, output_txt_path)
        elif ext in IMAGE_EXTENSIONS:
            ocr_image(local_destination_path, output_txt_path)
        elif ext in DOC_EXTENSIONS:
            extract_text_from_docx(local_destination_path, output_txt_path)
        elif ext == '.xlsx':
            extract_text_from_xlsx(local_destination_path, output_txt_path)
        elif ext == '.xls':
            extract_text_from_xls(local_destination_path, output_txt_path)
        else:
            print(f"Tipo no soportado: {local_destination_path}")

# --- MAIN ---

def main():
    print(f"Modo: {SOURCE_TYPE}")

    if SOURCE_TYPE == "LOCAL":
        files = list_files_local(LOCAL_SOURCE_DIR)
        process_files(files, DESTINATION_DIR, LOCAL_SOURCE_DIR)

    else:
        print("⚠️ FTP deshabilitado en esta versión final.")

    print("✅ Proceso finalizado.")

if __name__ == "__main__":
    os.makedirs(DESTINATION_DIR, exist_ok=True)
    main()
